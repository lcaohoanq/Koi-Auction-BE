import pandas as pd
import xml.etree.ElementTree as ET
import glob
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

# Function to convert XML to DataFrame
def xml_to_dataframe(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()
    all_records = []
    for testcase in root.findall('testcase'):
        record = {}
        record['name'] = testcase.get('name')
        record['classname'] = testcase.get('classname')
        record['time'] = float(testcase.get('time'))
        failure = testcase.find('failure')
        record['status'] = 'Failed' if failure is not None else 'Passed'
        all_records.append(record)
    return pd.DataFrame(all_records)

# Convert all XML files in the target/surefire-reports directory
xml_files = glob.glob('target/surefire-reports/*.xml')
all_dfs = [xml_to_dataframe(xml_file) for xml_file in xml_files]
final_df = pd.concat(all_dfs, ignore_index=True)

# Save to Excel
excel_path = 'target/surefire-reports/test-report.xlsx'
final_df.to_excel(excel_path, index=False)

# Load the workbook and select the active worksheet
wb = load_workbook(excel_path)
ws = wb.active
ws.title = "Detailed Test Results"

# Define styles
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
passed_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
failed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Apply formatting to header
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = alignment
    cell.border = border

# Format data cells and apply conditional formatting
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = border
        cell.alignment = alignment
        if cell.column == 4:  # Status column
            if cell.value == 'Passed':
                cell.fill = passed_fill
            elif cell.value == 'Failed':
                cell.fill = failed_fill

# Adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = min(max_length + 2, 50)  # Cap width at 50 characters
    ws.column_dimensions[column].width = adjusted_width

# Create summary sheet
summary_ws = wb.create_sheet("Test Summary")
summary_ws.append(["Class", "Total Tests", "Passed", "Failed"])

# Prepare data for class-wise charts
class_data = final_df.groupby('classname').agg({
    'name': 'count',
    'status': lambda x: (x == 'Passed').sum()
}).reset_index()
class_data.columns = ['Class', 'Total Tests', 'Passed']
class_data['Failed'] = class_data['Total Tests'] - class_data['Passed']

# Add data to summary sheet
for _, row in class_data.iterrows():
    summary_ws.append(row.tolist())

# Format summary sheet
for cell in summary_ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = alignment

# Adjust column widths in summary sheet
for col in summary_ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    summary_ws.column_dimensions[column].width = adjusted_width

# Create charts for each class
chart_row = 2
for idx, row in enumerate(summary_ws.iter_rows(min_row=2, values_only=True), start=2):
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = f"Test Results for {row[0]}"
    chart.y_axis.title = 'Number of Tests'
    chart.x_axis.title = 'Status'

    data = Reference(summary_ws, min_col=3, min_row=idx, max_col=4, max_row=idx)
    cats = Reference(summary_ws, min_col=3, min_row=1, max_col=4, max_row=1)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)

    chart_cell = f"F{chart_row}"
    summary_ws.add_chart(chart, chart_cell)
    
    # Move to the next position for the next chart
    chart_row += 15  # Adjust this value to increase/decrease space between charts

# Save the workbook
wb.save(excel_path)

print(f"Excel report with enhanced formatting and separate charts generated: {excel_path}")
